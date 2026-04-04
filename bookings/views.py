from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.core.mail import send_mail
from django.core.mail import EmailMultiAlternatives
from django.utils.html import strip_tags
from django.conf import settings
from django.utils import timezone
from .models import Booking, Review, SiteVisit, SERVICE_CHOICES
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import io, json
from datetime import datetime, date, timedelta
from django.db.models import Count
from django.db.models.functions import TruncDate, TruncMonth, TruncYear
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect
from django.db import IntegrityError
from django.core.mail import send_mail
from django.conf import settings
from django.http import HttpResponse
from django.db import IntegrityError



    
def get_booked_slots(request):
    date = request.GET.get('date')

    if not date:
        return JsonResponse({'slots': []})

    bookings = Booking.objects.filter(preferred_date=date)
    booked_times = [b.preferred_time.strftime('%H:%M') for b in bookings]

    return JsonResponse({'slots': booked_times})


print("🔥 EMAIL FUNCTION TRIGGERED")
# ── label maps ────────────────────────────────────────────────────────────────
SVC_MAP      = dict(SERVICE_CHOICES)
DUR_MAP      = {'30':'30 min','60':'60 min','90':'90 min','120':'120 min'}
# GENDER_MAP   = {'male':'Male','female':'Female','prefer_not':'Prefer not to say'}
THERAPY_MAP  = {'any':'No preference','female':'Female therapist','male':'Male therapist'}
PRESSURE_MAP = {'light':'Light','medium':'Medium','firm':'Firm','deep':'Deep'}



def update_booking(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)

    if request.method == "POST":
        action = request.POST.get("action")

        # ✅ CANCEL
        if action == "cancel":
            booking.status = "cancelled"
            booking.save()

            html_content = f"""
            <div style="font-family:DM Sans; background:#F5EFE0; padding:20px;">
              <div style="max-width:500px;margin:auto;background:#FFFDF8;padding:25px;border-radius:12px;">
                <h2 style="color:#5C4A32;">Salford & Co.</h2>

                <p>Hi <b>{booking.full_name}</b>,</p>

                <p>Your booking has been 
                <span style="color:#C0392B;"><b>cancelled</b></span>.</p>

                <div style="margin-top:15px;background:#FDEDEC;padding:12px;border-radius:8px;">
                  <b>Date:</b> {booking.preferred_date}<br>
                  <b>Time:</b> {booking.preferred_time}
                </div>

                <p style="margin-top:15px;">We apologize for any inconvenience.</p>
              </div>
            </div>
            """

        # ✅ POSTPONE
        elif action == "postpone":
            new_date = request.POST.get("new_date")
            new_time = request.POST.get("new_time")

            booking.preferred_date = new_date
            booking.preferred_time = new_time
            booking.status = "postponed"
            booking.save()

            html_content = f"""
            <div style="font-family:DM Sans; background:#F5EFE0; padding:20px;">
              <div style="max-width:500px;margin:auto;background:#FFFDF8;padding:25px;border-radius:12px;">
                <h2 style="color:#5C4A32;">Salford & Co.</h2>

                <p>Hi <b>{booking.full_name}</b>,</p>

                <p>Your booking has been 
                <span style="color:#8B7355;"><b>rescheduled</b></span>.</p>

                <div style="margin-top:15px;background:#EDE3CF;padding:12px;border-radius:8px;">
                  <b>New Date:</b> {new_date}<br>
                  <b>New Time:</b> {new_time}
                </div>

                <p style="margin-top:15px;">We look forward to serving you ✨</p>
              </div>
            </div>
            """

        # ✅ SEND EMAIL
        subject = "Booking Update - Salford & Co."
        text_content = strip_tags(html_content)

        email = EmailMultiAlternatives(
            subject,
            text_content,
            settings.EMAIL_HOST_USER,
            [booking.email],
        )
        email.attach_alternative(html_content, "text/html")
        email.send()

    return redirect('/dashboard/')

def _get_client_ip(request):
    x_forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded:
        return x_forwarded.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


from django.core.mail import EmailMultiAlternatives
from django.utils.html import strip_tags

def send_booking_email(booking):
    print("🔥 EMAIL FUNCTION TRIGGERED")

    svc   = SVC_MAP.get(booking.service, booking.service)
    dur   = DUR_MAP.get(booking.duration, booking.duration)
    date_ = booking.preferred_date.strftime('%d %B %Y')
    time_ = booking.preferred_time.strftime('%I:%M %p')

    # ───── OWNER EMAIL (PREMIUM) ─────
    owner_html = f"""
    <div style="font-family:DM Sans; background:#F5EFE0; padding:20px;">
      <div style="background:#FFFDF8; padding:25px; border-radius:12px;">
        
        <h2 style="color:#5C4A32;">🧾 New Booking Received</h2>

        <p><b>Name:</b> {booking.full_name}</p>
        <p><b>Email:</b> {booking.email}</p>
        <p><b>Phone:</b> {booking.phone}</p>

        <hr>

        <div style="background:#EDE3CF;padding:12px;border-radius:8px;">
          <b>Service:</b> {svc}<br>
          <b>Duration:</b> {dur}<br>
          <b>Date:</b> {date_}<br>
          <b>Time:</b> {time_}
        </div>

      </div>
    </div>
    """

    msg = EmailMultiAlternatives(
        subject=f"New Booking: {booking.full_name}",
        body=strip_tags(owner_html),
        from_email=settings.EMAIL_HOST_USER,
        to=[settings.OWNER_EMAIL],
    )
    msg.attach_alternative(owner_html, "text/html")
    msg.send()

    # ───── CLIENT EMAIL (PREMIUM UI) ─────
    client_html = f"""
    <div style="font-family:DM Sans; background:#F5EFE0; padding:20px;">
      <div style="max-width:500px;margin:auto;background:#FFFDF8;padding:25px;border-radius:12px;">
        
        <h2 style="font-family:Playfair Display;color:#5C4A32;">
          Salford & Co.
        </h2>

        <p>Hi <b>{booking.full_name}</b>,</p>

        <p>Your booking has been 
        <span style="color:#5A8A6A;"><b>confirmed</b></span> ✨</p>

        <div style="margin-top:15px;background:#EDE3CF;padding:12px;border-radius:8px;">
          <b>Service:</b> {svc}<br>
          <b>Date:</b> {date_}<br>
          <b>Time:</b> {time_}
        </div>

        <p style="margin-top:15px;">
          We look forward to giving you a relaxing experience 🌿
        </p>

        <p style="font-size:12px;color:#9B8B78;">
          Salford & Co. • Luxury Wellness Experience
        </p>

      </div>
    </div>
    """

    msg2 = EmailMultiAlternatives(
        subject="Your Booking is Confirmed ✨",
        body=strip_tags(client_html),
        from_email=settings.EMAIL_HOST_USER,
        to=[booking.email],
    )
    msg2.attach_alternative(client_html, "text/html")
    msg2.send()

# ── Main booking page ─────────────────────────────────────────────────────────
from django.db import IntegrityError

def booking_form(request):
    # Record visit
    SiteVisit.objects.create(
        ip_address=_get_client_ip(request),
        user_agent=request.META.get('HTTP_USER_AGENT', '')[:300],
    )

    reviews = Review.objects.filter(approved=True).order_by('-created_at')[:12]

    if request.method == 'POST':
        data = request.POST

        try:
            # ── Parse date/time ──
            raw_dob  = (data.get('date_of_birth') or '').strip()
            raw_date = (data.get('preferred_date') or '').strip()
            raw_time = (data.get('preferred_time') or '').strip()

            if not raw_date:
                return JsonResponse({'status': 'error', 'message': 'Preferred date is required.'}, status=400)

            if not raw_time:
                return JsonResponse({'status': 'error', 'message': 'Preferred time is required.'}, status=400)

            parsed_date = datetime.strptime(raw_date, '%Y-%m-%d').date()
            parsed_time = datetime.strptime(raw_time, '%H:%M').time()
            parsed_dob  = datetime.strptime(raw_dob, '%Y-%m-%d').date() if raw_dob else None

            # 🚫 STEP 1: Pre-check (fast UX)
            if Booking.objects.filter(
                preferred_date=parsed_date,
                preferred_time=parsed_time
            ).exists():
                return JsonResponse({
                    'status': 'error',
                    'message': '❌ This time slot is already booked. Please choose another time.'
                }, status=400)

            # 🚫 STEP 2: DB-safe create
            try:
                if data.get('client_type') == 'couple':
                    if not data.get('partner_name'):
                        return JsonResponse({
                            'status': 'error',
                            'message': 'Partner name is required for couple booking.'
                        }, status=400)
                booking = Booking.objects.create(
                    full_name=data.get('full_name','').strip(),
                    email=data.get('email','').strip(),
                    phone=data.get('phone','').strip(),
                    date_of_birth=parsed_dob,
                    partner_name=data.get('partner_name') or None,
                    partner_age=data.get('partner_age') or None,
                    partner_gender=data.get('partner_gender') or None,
                    # gender=data.get('gender','prefer_not'),
                    service=data.get('service','foot_massage'),
                    duration=data.get('duration','60'),
                    preferred_date=parsed_date,
                    preferred_time=parsed_time,
                    # therapist_preference=data.get('therapist_preference','any'),
                    health_conditions=data.get('health_conditions','').strip() or None,
                    allergies=data.get('allergies','').strip() or None,
                    pressure_preference=data.get('pressure_preference','medium'),
                    special_requests=data.get('special_requests','').strip() or None,
                    consent_given=data.get('consent_given') == 'on',
                )

            except IntegrityError:
                return JsonResponse({
                    'status': 'error',
                    'message': '❌ This slot was just booked by someone else. Try another time.'
                }, status=400)

            # ✅ Send email after success
            send_booking_email(booking)

            return JsonResponse({
                'status': 'success',
                'name': booking.full_name,
                'booking_id': booking.id
            })

        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=400)

    return render(request, 'bookings/booking_form.html', {'reviews': reviews})


# ── Review submit ─────────────────────────────────────────────────────────────
def submit_review(request):
    if request.method == 'POST':
        data = request.POST
        try:
            Review.objects.create(
                name=data.get('name','').strip(),
                service=data.get('service','foot_massage'),
                rating=int(data.get('rating', 5)),
                comment=data.get('comment','').strip(),
            )
            return JsonResponse({'status':'success'})
        except Exception as e:
            return JsonResponse({'status':'error','message':str(e)}, status=400)
    return JsonResponse({'status':'error'}, status=405)


# ── Admin dashboard with analytics ───────────────────────────────────────────
@login_required(login_url='/login/')
def admin_dashboard(request):
    if request.META.get('REMOTE_ADDR') != '192.168.1.53':
        return HttpResponse("Unauthorized", status=403)
    bookings = Booking.objects.all().order_by('-created_at')
    reviews  = Review.objects.all().order_by('-created_at')
    now      = timezone.now()
    today    = now.date()

    # ── Visitor stats ──────────────────────────────────────────────────────
    visits_today   = SiteVisit.objects.filter(visited_at__date=today).count()
    visits_month   = SiteVisit.objects.filter(
        visited_at__year=today.year, visited_at__month=today.month).count()
    visits_year    = SiteVisit.objects.filter(visited_at__year=today.year).count()
    visits_total   = SiteVisit.objects.count()

    # Last 30 days — daily breakdown
    thirty_ago = today - timedelta(days=29)
    daily_qs = (
        SiteVisit.objects
        .filter(visited_at__date__gte=thirty_ago)
        .annotate(day=TruncDate('visited_at'))
        .values('day')
        .annotate(count=Count('id'))
        .order_by('day')
    )
    # Fill gaps with 0
    daily_map = {entry['day']: entry['count'] for entry in daily_qs}
    daily_labels, daily_data = [], []
    for i in range(30):
        d = thirty_ago + timedelta(days=i)
        daily_labels.append(d.strftime('%d %b'))
        daily_data.append(daily_map.get(d, 0))

    # Last 12 months
    monthly_qs = (
        SiteVisit.objects
        .filter(visited_at__year__gte=today.year-1)
        .annotate(month=TruncMonth('visited_at'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )
    monthly_labels = [e['month'].strftime('%b %Y') for e in monthly_qs]
    monthly_data   = [e['count'] for e in monthly_qs]

    # Per-year
    yearly_qs = (
        SiteVisit.objects
        .annotate(yr=TruncYear('visited_at'))
        .values('yr')
        .annotate(count=Count('id'))
        .order_by('yr')
    )
    yearly_labels = [e['yr'].strftime('%Y') for e in yearly_qs]
    yearly_data   = [e['count'] for e in yearly_qs]

    # Booking stats
    bookings_today = Booking.objects.filter(created_at__date=today).count()
    bookings_month = Booking.objects.filter(
        created_at__year=today.year, created_at__month=today.month).count()

    # Service breakdown for chart
    svc_counts = {}
    for b in bookings:
        k = SVC_MAP.get(b.service, b.service)
        svc_counts[k] = svc_counts.get(k, 0) + 1

    context = {
        'bookings': bookings,
        'reviews':  reviews,
        # visit stats
        'visits_today':  visits_today,
        'visits_month':  visits_month,
        'visits_year':   visits_year,
        'visits_total':  visits_total,
        # chart data (JSON for JS)
        'daily_labels':   json.dumps(daily_labels),
        'daily_data':     json.dumps(daily_data),
        'monthly_labels': json.dumps(monthly_labels),
        'monthly_data':   json.dumps(monthly_data),
        'yearly_labels':  json.dumps(yearly_labels),
        'yearly_data':    json.dumps(yearly_data),
        # booking stats
        'bookings_today': bookings_today,
        'bookings_month': bookings_month,
        'svc_labels':     json.dumps(list(svc_counts.keys())),
        'svc_data':       json.dumps(list(svc_counts.values())),
    }
    return render(request, 'bookings/admin_dashboard.html', context)


# ── Excel export ──────────────────────────────────────────────────────────────
def export_excel(request):
    bookings = Booking.objects.all().order_by('-created_at')

    BEIGE  = "F5EFE0"; TAUPE  = "8B7355"; DARK   = "2C2416"
    WHITE  = "FFFFFF"; ALT    = "EDE3CF"; ACCENT = "C8B89A"
    GREEN  = "5A8A6A"; RED    = "C0392B"

    wb = openpyxl.Workbook()

    def hdr(ws, row, col, val, bg=TAUPE, fg=WHITE, sz=11):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name='Calibri', size=sz, bold=True, color=fg)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = Border(bottom=Side(style='medium', color=DARK),
                             right=Side(style='thin', color=ACCENT))
        return c

    def dat(ws, row, col, val, bg=BEIGE, center=False):
        c = ws.cell(row=row, column=col, value=val)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.font      = Font(name='Calibri', size=10)
        c.alignment = Alignment(vertical='center',
                                horizontal='center' if center else 'left',
                                wrap_text=True)
        c.border    = Border(bottom=Side(style='thin', color=ACCENT),
                             right=Side(style='thin', color=ACCENT))
        return c

    # Sheet 1 — All Bookings
    ws = wb.active
    ws.title = "All Bookings"
    ws.merge_cells('A1:P1')
    t = ws['A1']
    t.value     = "SALFORD & CO. — RELAXING SPA  |  Client Booking Register 2026"
    t.font      = Font(name='Georgia', size=16, bold=True, color=WHITE)
    t.fill      = PatternFill("solid", fgColor=DARK)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 38

    ws.merge_cells('A2:P2')
    s = ws['A2']
    s.value     = (f"Delhi NCR  •  8796044169  •  rrony231688@gmail.com  •  "
                   f"Generated {datetime.now().strftime('%d %B %Y, %I:%M %p')}")
    s.font      = Font(name='Calibri', size=9, italic=True, color=ACCENT)
    s.fill      = PatternFill("solid", fgColor="3A2E1E")
    s.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8

    headers = ["#","Full Name","Email","Phone","Service","Duration",
               "Date","Time","Therapist","Health Cond.","Allergies",
               "Pressure","Special Requests","Consent","Booked At"]
    widths  = [4,20,26,15,12,20,12,13,11,16,22,18,12,26,9,20]
    for ci,(h,w) in enumerate(zip(headers,widths),1):
        hdr(ws,4,ci,h)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[4].height = 28
    ws.freeze_panes = ws['A5']

    for ri,b in enumerate(bookings,1):
        row = ri+4; bg = BEIGE if ri%2==1 else ALT
        vals = [
            ri, b.full_name, b.email, b.phone,
            b.client_type,
            SVC_MAP.get(b.service,b.service),
            DUR_MAP.get(b.duration,b.duration),
            b.preferred_date.strftime('%d/%m/%Y') if b.preferred_date else '—',
            b.preferred_time.strftime('%I:%M %p')  if b.preferred_time else '—',
            # THERAPY_MAP.get(b.therapist_preference,b.therapist_preference),
            b.health_conditions or '—', b.allergies or '—',
            PRESSURE_MAP.get(b.pressure_preference,b.pressure_preference),
            b.special_requests or '—',
            '✔ Yes' if b.consent_given else '✘ No',
            b.created_at.strftime('%d/%m/%Y %I:%M %p'),
        ]
        for ci,val in enumerate(vals,1):
            c = dat(ws,row,ci,val,bg,center=(ci in [1,5,7,9,10,13,15]))
            if ci==15:
                c.font = Font(name='Calibri',size=10,bold=True,
                              color=GREEN if b.consent_given else RED)
        ws.row_dimensions[row].height = 20

    fr = len(bookings)+6
    ws.merge_cells(f'A{fr}:D{fr}')
    fc=ws[f'A{fr}']; fc.value=f"Total Bookings: {len(bookings)}"
    fc.font=Font(name='Calibri',size=11,bold=True,color=DARK)
    fc.fill=PatternFill("solid",fgColor=ACCENT)
    fc.alignment=Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[fr].height=22

    # Sheet 2 — Service Summary
    ws2=wb.create_sheet("Service Summary")
    ws2.merge_cells('A1:E1')
    ws2['A1'].value="BOOKINGS BY SERVICE"
    ws2['A1'].font=Font(name='Georgia',size=14,bold=True,color=WHITE)
    ws2['A1'].fill=PatternFill("solid",fgColor=DARK)
    ws2['A1'].alignment=Alignment(horizontal='center',vertical='center')
    ws2.row_dimensions[1].height=32
    counts={}
    for b in bookings:
        k=SVC_MAP.get(b.service,b.service); counts[k]=counts.get(k,0)+1
    hdr(ws2,3,1,"Service",TAUPE); hdr(ws2,3,2,"Bookings",TAUPE)
    ws2.column_dimensions['A'].width=26; ws2.column_dimensions['B'].width=14
    for i,(svc,cnt) in enumerate(sorted(counts.items(),key=lambda x:-x[1]),4):
        bg=BEIGE if i%2==0 else ALT
        dat(ws2,i,1,svc,bg); dat(ws2,i,2,cnt,bg,center=True)
    if counts:
        chart=BarChart(); chart.type="col"; chart.title="Bookings by Service"
        chart.y_axis.title="Count"; chart.style=10; chart.width=18; chart.height=11
        dr=Reference(ws2,min_col=2,min_row=3,max_row=3+len(counts))
        cr=Reference(ws2,min_col=1,min_row=4,max_row=3+len(counts))
        chart.add_data(dr,titles_from_data=True); chart.set_categories(cr)
        ws2.add_chart(chart,"D3")

    # Sheet 3 — Reviews
    ws3=wb.create_sheet("Client Reviews")
    ws3.merge_cells('A1:F1')
    ws3['A1'].value="CLIENT REVIEWS"
    ws3['A1'].font=Font(name='Georgia',size=14,bold=True,color=WHITE)
    ws3['A1'].fill=PatternFill("solid",fgColor=DARK)
    ws3['A1'].alignment=Alignment(horizontal='center',vertical='center')
    ws3.row_dimensions[1].height=32
    for ci,(h,w) in enumerate(zip(["#","Name","Service","Rating","Comment","Date"],
                                   [4,20,22,10,40,18]),1):
        hdr(ws3,3,ci,h,TAUPE); ws3.column_dimensions[get_column_letter(ci)].width=w
    for ri,rv in enumerate(Review.objects.all().order_by('-created_at'),1):
        row=ri+3; bg=BEIGE if ri%2==1 else ALT
        stars='★'*rv.rating+'☆'*(5-rv.rating)
        dat(ws3,row,1,ri,bg,center=True); dat(ws3,row,2,rv.name,bg)
        dat(ws3,row,3,SVC_MAP.get(rv.service,rv.service),bg)
        c=dat(ws3,row,4,stars,bg,center=True)
        c.font=Font(name='Calibri',size=11,bold=True,
                    color='C8920A' if rv.rating>=4 else TAUPE)
        dat(ws3,row,5,rv.comment,bg)
        dat(ws3,row,6,rv.created_at.strftime('%d/%m/%Y'),bg,center=True)
        ws3.row_dimensions[row].height=22

    # Sheet 4 — Contact List
    ws4=wb.create_sheet("Contact List")
    ws4.merge_cells('A1:E1')
    ws4['A1'].value="CLIENT CONTACT LIST"
    ws4['A1'].font=Font(name='Georgia',size=14,bold=True,color=WHITE)
    ws4['A1'].fill=PatternFill("solid",fgColor=DARK)
    ws4['A1'].alignment=Alignment(horizontal='center',vertical='center')
    ws4.row_dimensions[1].height=30
    for ci,(h,w) in enumerate(zip(["Full Name","Email","Phone","Service","Appt. Date"],
                                   [22,28,16,22,16]),1):
        hdr(ws4,3,ci,h,TAUPE); ws4.column_dimensions[get_column_letter(ci)].width=w
    for ri,b in enumerate(bookings,1):
        row=ri+3; bg=BEIGE if ri%2==1 else ALT
        dat(ws4,row,1,b.full_name,bg); dat(ws4,row,2,b.email,bg)
        dat(ws4,row,3,b.phone,bg)
        dat(ws4,row,4,SVC_MAP.get(b.service,b.service),bg)
        dat(ws4,row,5,b.preferred_date.strftime('%d/%m/%Y') if b.preferred_date else '—',bg,center=True)

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    fname=f"SalfordSpa_Bookings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    resp=HttpResponse(buf.getvalue(),
         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition']=f'attachment; filename="{fname}"'
    return resp
